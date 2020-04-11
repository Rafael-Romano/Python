from openpyxl import Workbook
import calendar as cld
# import pandas as pd
import numpy as np

print("FICHA DE FREQUÊNCIA!")
mes = int(input('Informe o mês: '))
ano = int(input('Informe o ano: '))
dt = np.array(cld.month(ano, mes))

semana = [
    ("Nome", "Sexo", "Idade"),
    ("Rafa", "M", 28),
    ("Mauricio", "M", 24),
    ("Alex", "M", 19),
    ("Claudiana", "F", 44),
    ("Amauri", "M", 55)
]

arquivo_excel = Workbook()
planilha1 = arquivo_excel.active
planilha1.title = "Ficha de frequência"

for l in semana:
    planilha1.append(l)



max_linha = planilha1.max_row
max_coluna = planilha1.max_column

for l in semana:
    planilha1.append(l)
for i in range(1, max_linha + 1):
    for j in range(1, max_coluna + 1):
        print(planilha1.cell(row=i, column=j).value, end=' ')

# def checar_horas():
# def escrever():

# ler = pd.read_excel(r'C:\Users\roman\Documents\Ficha_teste.xlsx')
# print(semana['SEG'][1])
# print(arquivo_excel.sheetnames)
